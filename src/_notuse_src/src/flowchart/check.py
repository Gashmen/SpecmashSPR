#Cell fill color fgColor='FFF8F2D8'
import openpyxl
from openpyxl.styles import PatternFill
def get_cell_values(sheet):

    cell_values = {}
    for row in sheet.iter_rows():
        for cell in row:
            # Получение значения ячейки
            value = cell.value
            # Получение координат ячейки
            coordinates = cell.coordinate
            # Добавление значения и его координат в словарь
            cell_values[coordinates] = value

    return cell_values

class ActionTechProccess:

    code='005'#Первый код номер

    def __init__(self):
        '''Первая строчка'''
        self.workshop = None #Цех номер
        self.section = None#участок
        self.RM = None#РМ

        self.code = ActionTechProccess.code
        self.create_new_code_number()

        self.operation_name_main = None#Код,наименование операции
        self.iot_number = None#Номер ИОТ
        '''Вторая строчка'''
        self.equipment_name_main = None#Оборудование, котором осуществляется тех операция
        self.CM = None#СМ столбец
        self.profile = None #Проф. столбей
        self.R = None#Р столбец
        self.UT = None#УТ столбец
        self.KR = None# КР столбец
        self.KOID = None# КОИД столбец
        self.EN = None# ЕН столбец
        self.OP = None#ОП столбец
        self.K_sht = None#Кшт столбец
        self.T_pz = None#Тп.з. столбец
        self.T_sht = None#Тшт столбец

    def set_first_row(self,
                      workshop=None,
                      section=None,
                      RM=None,
                      operation_name_main=None,
                      iot_number=None):
        '''установка значений первой строке'''
        self.workshop = workshop
        self.section=section
        self.RM=RM
        self.operation_name_main=operation_name_main
        self.iot_number=iot_number

    def set_second_row(self,
                       equipment_name_main=None,
                       CM=None,
                       profile=None,
                       R=None,
                       UT=None,
                       KR=None,
                       KOID=None,
                       EN=None,
                       OP=None,
                       K_sht=None,
                       T_pz=None,
                       T_sht=None):
        self.equipment_name_main = equipment_name_main  # Оборудование, котором осуществляется тех операция
        self.CM = CM  # СМ столбец
        self.profile = profile  # Проф. столбей
        self.R = R  # Р столбец
        self.UT = UT  # УТ столбец
        self.KR = KR  # КР столбец
        self.KOID = KOID  # КОИД столбец
        self.EN = EN  # ЕН столбец
        self.OP = OP  # ОП столбец
        self.K_sht = K_sht  # Кшт столбец
        self.T_pz = T_pz  # Тп.з. столбец
        self.T_sht = T_sht  # Тшт столбец

    def create_new_code_number(self):
        '''
        Задание нового код номера технологического процесса, чтобы при вызове функции автоматически прибавлялось 005
        :return:
        '''
        new_code = str(int(ActionTechProccess) + 5)
        if len(new_code) == 1:
            ActionTechProccess.code = '00' + new_code
        elif len(new_code) == 2:
            ActionTechProccess.code = '0' + new_code
        else:
            ActionTechProccess.code = new_code



# def create_new_main_row(N_cell='',X_cell='', AR_cell='005',BF_cell='',current_cell=83,sheet_number=1, )


# Вывод значений ячеек и их координат

if __name__ == "__main__":



    workbook = openpyxl.load_workbook('шаблон.xlsx')

    ws = workbook.active

    ws['IF32'].value = '1234'
    cell_fill = '#' + ws['CS32'].fill.start_color.index
    cell_fill_next = '#' + ws['IF32'].fill.start_color.index
    print(cell_fill,cell_fill_next)
    ws['IF32'].fill = PatternFill('solid', fgColor='FFF8F2D8')
    workbook.save('шаблон_тест.xlsx')





    # print(cell_values)
