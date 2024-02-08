import os
import sys

import openpyxl


from PyQt5 import QtCore, QtGui, QtWidgets
# from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QMessageBox, QDialog

from src.authentication import frontend_auth

from src.qt_creating import main_ui
from src.qt_creating import start_ui
from src.scripts import scripts_start
from src.smbconnect import smbconnect

import logging
import tempfile

class AuthWindow(QtWidgets.QMainWindow, frontend_auth.Ui_WelcomeWindow):

    def __init__(self):
        '''ИНИЦИАЛИЗАЦИЯ'''

        super().__init__()

        self.setupUi(self)
        self.set_logger()#Установка логгера
        self.create_connect_SMB()#Установка соединения с сервером
        self.get_verification_path()

        self.write_username()

        '''PUSH BUTTONS ACTION'''
        self.junction_boxes_pushButton.clicked.connect(self.show_jb_qt)
        self.scripts_pushButton.clicked.connect(self.show_window_scripts)

    def set_logger(self):
        try:
            self.logger_file_obj = tempfile.NamedTemporaryFile(delete=False)#Создание временного файла
            self.logger_path = self.logger_file_obj.name #Получение пути данного временного файла
            #Настройка логгера
            self.logger = logging.getLogger(__name__)
            self.logger.setLevel(logging.DEBUG)
            self.logger_file = logging.FileHandler(filename=self.logger_path)
            self.logger_file.setFormatter(logging.Formatter(fmt='[%(asctime)s: %(levelname)s] %(name)s %(message)s'))
            self.logger.addHandler(self.logger_file)
            self.logger_file_obj.close()
            self.logger.info('Логгер установлен')
        except Exception as e:
            raise e

    def create_connect_SMB(self):
        try:
            self.smb_connect = smbconnect.SMBCONNECT_SPECMASH_SERVER()
            self.smb_connect.install_connect()
        except Exception as e:
            self.logger.debug('Соединенние с сервером по SMB не установленно')

    def get_verification_path(self):
        self.smb_connect.get_auth_path()
        try:
            self.path_to_bd_xlsx = self.smb_connect.authentication_path
            if not self.path_to_bd_xlsx.endswith('.xlsx'):
                os.rename(src=self.path_to_bd_xlsx,
                          dst=self.path_to_bd_xlsx + '.xlsx')
                self.path_to_bd_xlsx = self.path_to_bd_xlsx + '.xlsx'
            self.logger.info(f'База данных сотрудников определена {self.path_to_bd_xlsx}')
        except:
            self.path_to_bd_xlsx = ''
            self.logger.debug(f'База данных сотрудников не прогрузилась')

    def create_dict_for_verification(self):
        '''
        Создает словарь для верификации, определяет имя и почту
        :return:
        '''

        if self.path_to_bd_xlsx != '':
            wb = openpyxl.load_workbook(self.path_to_bd_xlsx)
            ws = wb.active
            self.dict_first_second_name = dict()
            for cell in ws['F']:
                if cell.value:
                    if '@' in cell.value:
                        value = cell.value.split('@')[0]
                        full_name = ws[f'B{cell.row}'].value
                        first_name = full_name.split(' ')[0]
                        second_name = full_name.split(' ')[1]
                        third_name = None
                        if len(full_name.split(' ')) >= 3:
                            third_name = full_name.split(' ')[2]
                        if third_name:
                            self.dict_first_second_name[value] = first_name + f' {second_name[0]}.{third_name[0]}.'
                        else:
                            self.dict_first_second_name[value] = first_name + f' {second_name[0]}.'
            if self.dict_first_second_name != dict():
                self.logger.info('Получена информация про логины и почты сотрудников')
            else:
                self.logger.debug('Информация по сотрудникам офиса СПБ отсутствует')


    def write_username(self):
        '''
        Записываем кто разработал
        :return:
        '''
        self.create_dict_for_verification()

        if (os.getlogin() != '' and os.getlogin() != 'admin'):
            self.logger.info(f'Инициализируется пользователь с логином {os.getlogin()}')
            if os.getlogin() in list(self.dict_first_second_name.keys()):
                self.username_QLineEdit.insert(self.dict_first_second_name[os.getlogin()])
            else:
                self.logger.debug(f'{os.getlogin()} не входит в список сотрудников')


    def return_username(self):
        '''Возвращает username'''
        return self.username_QLineEdit.text()

    def return_task_number(self):
        '''Возвращает номер заявки'''
        return self.number_task_QLineEdit.text()

    def return_position_number(self):
        '''Возвращает номер заявки'''
        return self.position_number_QLineEdit.text()

    def set_task_number(self,task_number):
        '''Устанавливает номер задачи'''
        self.number_task_QLineEdit.setText(task_number)

    def set_position_number(self,position_number):
        '''Устанавливает номер задачи'''
        self.position_number_QLineEdit.setText(position_number)

    def show_jb_qt(self):

        path_to_csv = '\\'.join(os.getcwd().split('\\')[0:]) + '\\src\\bd'

        self.smb_connect.get_doc_path()
        path_to_dxf_shell = self.smb_connect.path
        path_to_terminal_dxf = self.smb_connect.path
        self.jb_window = main_ui.DxfCreator(path_to_csv=path_to_csv,
                                            path_to_dxf=path_to_dxf_shell,
                                            path_to_terminal_dxf=path_to_terminal_dxf)
        self.jb_window.set_logger(logger=self.logger)
        # self.jb_window = start_ui.Mainver(path_to_terminal_dxf=path_to_terminal_dxf)
        self.jb_window.designer_name = self.return_username()
        self.jb_window.task_number = self.return_task_number()
        self.jb_window.position_number = self.return_position_number()
        self.jb_window.label.setText(f'{self.jb_window.designer_name}    {self.jb_window.task_number}')
        self.close()
        self.jb_window.show()

    def show_window_scripts(self):
        self.scripts_window = scripts_start.ScriptsMainWindow()
        self.close()
        self.scripts_window.show()

    def closeEvent(self, event):
        try:
            self.smb_connect.save_log(logger_path=self.logger_path)
        except:
            pass


def run_app():
    app = QtWidgets.QApplication(sys.argv)
    authWindow = AuthWindow()
    authWindow.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    authWindow = AuthWindow()
    authWindow.show()
    sys.exit(app.exec_())
