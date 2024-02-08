import math
import os

import ezdxf
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5 import QtCore, QtGui, QtWidgets

def create_doc_BOM(dxfbase_path:str):
    '''Создает BOM удаляя все не нужное'''
    doc_bom = ezdxf.readfile(dxfbase_path)

    doc_dxfbase_for_del = ezdxf.readfile(dxfbase_path)

    doc_bom.modelspace().delete_all_entities()

    blocks_BOM = ['BOM_FIRST','BOM_SECOND']


    for block in doc_dxfbase_for_del.blocks:
        try:
            if block.dxf.name not in blocks_BOM and '*' not in block.dxf.name:
                    doc_bom.blocks.delete_block(name=block.dxf.name)
        except:
            continue
    return doc_bom

def create_BOM_FIRST(doc_bom):
    '''Создает первый лист спецификации'''

    block_border = doc_bom.blocks['BOM_FIRST']
    values = {attdef.dxf.tag: '' for attdef in block_border.query('ATTDEF')}
    if doc_bom.blocks.get('BOM_FIRST'):
        border_insert = doc_bom.modelspace().add_blockref(name='BOM_FIRST',
                                                          insert=(0, 0))
        border_insert.add_auto_attribs(values)

        return border_insert

def create_BOM_SECOND(doc_bom):
    '''Создает первый лист спецификации'''

    block_border = doc_bom.blocks['BOM_SECOND']
    values = {attdef.dxf.tag: '' for attdef in block_border.query('ATTDEF')}
    if doc_bom.blocks.get('BOM_SECOND'):
        border_insert = doc_bom.modelspace().add_blockref(name='BOM_SECOND',
                                                          insert=(0, 0))
        border_insert.add_auto_attribs(values)

        return border_insert

def check_name_len(name_string:str):
    if name_string != None:
        if math.ceil(len(name_string)/27) ==1:
            return True
        else:
            return False

def read_BOM_base(xlsx_base_path:str):
    '''
    Чтение базавой эксельки BOM и выдача словаря
    :param xlsx_base_path:
    :return:
    '''

    return_dict = dict()

    workbook = openpyxl.load_workbook(xlsx_base_path)
    worksheet = workbook.active

    column_names = list(worksheet[1])

    for column_name in column_names:
        return_dict[column_name.value] = {}
        for count,row_value in enumerate(worksheet[get_column_letter(column_name.column)][column_name.row:]):
            return_dict[column_name.value][count] = row_value.value

    return return_dict

def write_attrib(BOM_insert,dict_for_writing_attrib):
    '''
    Заполение аттрибутов в спецификации
    :param BOM_insert: BOM_FIRST или BOM_SECOND
    :param dict_for_writing_attrib:
    {'Сборочные единицы':{0:{"Формат":"А4","Зона":"","Поз.":0+1,"Обозначение":"ВРПТ.301172.024-11","Наименование":"Оболочка ВП.161610", "Кол.":1, "Примечание": производитель}}
    :return:
    '''

    row_start = 2
    dict_name_attrib = {attrib.dxf.tag: attrib for attrib in BOM_insert.attribs}
    for main_type_name in dict_for_writing_attrib:
        # [Сборочные изделия, Детали, Стандартные изделия ...]
        dict_name_attrib[f'E{row_start}'].dxf.text = main_type_name
        row_start+=2

        dict_with_pozition = dict_for_writing_attrib[main_type_name]

        for number_pozition in dict_with_pozition:
            dict_name_attrib[f'A{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Формат']
            dict_name_attrib[f'B{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Зона']
            dict_name_attrib[f'C{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Поз.']
            dict_name_attrib[f'D{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Обозначение']
            dict_name_attrib[f'F{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Кол.']
            dict_name_attrib[f'G{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Примечание']

            if math.ceil(len(dict_with_pozition[number_pozition]['Наименование']) / 27) == 1:
                dict_name_attrib[f'E{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Наименование']
            elif math.ceil(len(dict_with_pozition[number_pozition]['Наименование']) / 27) == 2:
                dict_name_attrib[f'E{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Наименование'][:27]
                row_start+=1
                dict_name_attrib[f'E{row_start}'].dxf.text = dict_with_pozition[number_pozition]['Наименование'][27:]

    row_start+=1

def calculate_count_row_for_one_position(dict_with_all_info_in_BOM_row:dict)->int:
    '''
    Расчитать количество строк, которые понадобятся для этой позиции для заполнения в BOM
    :param dict_with_all_info_in_BOM_row: {'Обозначение': 'ВРПТ.301172.024-11', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм',
                                            'Формат': 'А4', 'Кол.': None, 'Примечание': None}
    :return: row:int
    '''
    row = 0
    for column_name, name_in_bom in dict_with_all_info_in_BOM_row.items():
        if name_in_bom== None:
            name_in_bom = ''
        name_in_bom = str(name_in_bom)
        if '#' not in name_in_bom:
            row = max(1, row)
        else:
            row = max(row, len(name_in_bom.split('#')))
    return row

def recreate_dict_with_all_info_in_BOM_row(dict_with_all_info_in_BOM_row:dict,row:int):
    '''
    Пересоздает словарь dict_with_all_info_in_BOM_row
    :param dict_with_all_info_in_BOM_row: {'Обозначение': 'ВРПТ.301172.024-11', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм',
                                            'Формат': 'А4', 'Кол.': None, 'Примечание': None}
    :param row: 4
    :return: {'Обозначение': ['ВРПТ.301172.024-11','',''], 'Наименование': ['Кабельный ввод ВЗ-Н25','для не бронированного','кабеля, диаметром 12-18мм'],
                'Формат': ['А4','',''], 'Кол.': ['','',''], 'Примечание': ['','','']}
    '''
    for column_name, name_in_bom in dict_with_all_info_in_BOM_row.items():
        if name_in_bom== None:
            name_in_bom = ''
        if '#' not in name_in_bom:
             for i in range(0, row):
                 if i == 0:
                     dict_with_all_info_in_BOM_row[column_name] = [name_in_bom]
                 else:
                     dict_with_all_info_in_BOM_row[column_name].append('')
        else:
            row_new = name_in_bom.split('#')
            for i in range(0, row):
                if len(row_new) == row:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                else:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                    for i in range(0,row - len(row_new)):
                        dict_with_all_info_in_BOM_row[column_name].append('')


def add_dict(dict_with_all_info_in_BOM_row:dict, count_row:int):
    '''
    Нужно добавить в этот же словарь тэг того, сколько страниц добавить в конце после прохода по данному словорю а также найти, сколько нужно добавить
    :param dict_with_all_info_in_BOM_row:{'Обозначение': 'ВРПТ.305311.001-025', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм', 'Свойство': 'Сборочные единицы', 'Формат': 'А4', 'Кол.': None, 'Примечание': None}
    :param count_row:1
    :return:
    '''
    #Получаем необходимо число строк
    row = calculate_count_row_for_one_position(dict_with_all_info_in_BOM_row)

    for column_name, name_in_bom in dict_with_all_info_in_BOM_row.items():
        if name_in_bom== None:
            name_in_bom = ''
        name_in_bom = str(name_in_bom)
        if '#' not in name_in_bom:
             for i in range(0, row):
                 if i == 0:
                     dict_with_all_info_in_BOM_row[column_name] = [name_in_bom]
                 else:
                     dict_with_all_info_in_BOM_row[column_name].append('')
        else:
            row_new = name_in_bom.split('#')
            for i in range(0, row):
                if len(row_new) == row:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                else:
                    dict_with_all_info_in_BOM_row[column_name] = row_new
                    for i in range(0,row - len(row_new)):
                        dict_with_all_info_in_BOM_row[column_name].append('')

    return row + count_row

def create_dict_main_properties(list_properties:list):
    '''
    Свойства нужно вынести наружу и сделать словарь
    :param list_properties:
[{'Обозначение': 'ВРПТ.301172.024-11', 'Наименование': 'Оболочка ВП.161610', 'Свойство': 'Сборочные единицы', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': None, 'Наименование': 'Винт А2.М6-6gx10.019#ГОСТ 17473-80', 'Свойство': 'Стандартные изделия', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': None, 'Наименование': 'Шайба 6 019 ГОСТ 6402-70', 'Свойство': 'Стандартные изделия', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': None, 'Наименование': 'Шайба A.6.019 ГОСТ 11371-78', 'Свойство': 'Стандартные изделия', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': 'ВРПТ.745551.005-140', 'Наименование': 'DIN-рейка NS35х7,5, L=140 мм', 'Свойство': 'Детали', 'Формат': 'А4', 'Кол.': None, 'Примечание': None},
 {'Обозначение': 'ВРПТ.305311.001-025', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм', 'Свойство': 'Сборочные единицы', 'Формат': 'А4', 'Кол.': None, 'Примечание': None}]

    :return: "Стандартные изделия":{}
    '''
    return_dict = dict()
    for equip_dict in list_properties:
        property = equip_dict['Свойство']
        if property not in return_dict:
            equip_dict.pop('Свойство')
            return_dict[property] = [equip_dict]
        else:
            equip_dict.pop('Свойство')
            return_dict[property].append(equip_dict)
    return return_dict

def create_list_all_block_names_in_doc(doc_new):
    '''
    Создает список всех блоков импортированных в self.doc_new
    :param doc: self.doc_new
    :return: ['SUPU_SCREW..., VP.161610_topside,...']
    '''
    return [block.dxf.name for block in doc.blocks if '*' not in block.dxf.name]


def check_next_page(BOM_insert_name:str, row_number:int):
    '''
    Проверка на создание следующей страницы
    :param BOM_insert_name: либо BOM_FIRST либо BOM_SECOND
    :param row_number: 1-29 или 1-32
    :return: True or False
    '''

    if BOM_insert_name == 'BOM_FIRST':
        if row_number > 29:
            return False
        else:
            return True
    elif BOM_insert_name == 'BOM_SECOND':
        if row_number > 32:
            return False
        else:
            return True

def write_mainproperty_in_bom_E_cell(BOM_insert_name:str, row_number:int, mainproperty_attribtag_name:str, dict_attribs:dict):
    '''
    Заполнение спецификации Сборочная единица, Деталь, и тд
    :return:True or False
    '''
    if check_next_page(BOM_insert_name=BOM_insert_name, row_number = row_number + 4):
        row_number +=1
        tag_attrib = 'E' + str(row_number)
        dict_attribs[tag_attrib].dxf.text = mainproperty_attribtag_name
        row_number += 2
        return True
    else:
        return False


def get_path_to_xlsx(main_window_class_instance):
    '''
    Получение пути для xlsx
    :param main_window_class_instance: параметр self
    :return:
    '''
    path_to_xlsx = QtWidgets.QFileDialog.getOpenFileName(main_window_class_instance,
                                                             directory='\\'.join(os.getcwd().split('\\')[0:-1]),
                                                             caption='Выбор файла бд xlsx ',
                                                             filter='Excel file(*.xlsx)')[0]



    return path_to_xlsx

def create_dict_with_insert_names(doc):
    '''
    doc того чертежа, откуда хотим создать BOM
    :param doc:
    :return: {имя инстера: количество его на чертеже}
    '''

    list_with_block_names = [insert.dxf.name for insert in doc.modelspace().query('INSERT') if
                             '*' not in insert.dxf.name]
    dict_with_block_names = {}

    for insert_name in list_with_block_names:
        if insert_name not in dict_with_block_names:
            dict_with_block_names[insert_name] = 1
        else:
            dict_with_block_names[insert_name] += 1

    return dict_with_block_names

def dict_all_attrib_in_BOM(list_for_creating_BOM_with):
    '''

    :param list_for_creating_BOM_with:
    {'Сборочные единицы': [{'Обозначение': 'ВРПТ.301172.024-021', 'Наименование': 'Оболочка ВП.262512', 'Формат': 'А4', 'Кол.': 1, 'Примечание': None, 'Цена': 7154.1}, {'Обозначение': 'ВРПТ.305311.001-025', 'Наименование': 'Кабельный ввод ВЗ-Н25#для не бронированного#кабеля, диаметром 12-18мм', 'Формат': 'А4', 'Кол.': 6, 'Примечание': None, 'Цена': 884.21}, {'Обозначение': 'ВРПТ.305311.001-016', 'Наименование': 'Кабельный ввод ВЗ-Н16#для не бронированного#кабеля, диаметром 3-8мм', 'Формат': 'А4', 'Кол.': 3, 'Примечание': None, 'Цена': 615.75}, {'Обозначение': 'ВРПТ.305311.001-012', 'Наименование': 'Кабельный ввод ВЗ-Н12#для не бронированного#кабеля, диаметром 2-6мм', 'Формат': 'А4', 'Кол.': 2, 'Примечание': None, 'Цена': 607.79}, {'Обозначение': 'ВРПТ.305311.001-032', 'Наименование': 'Кабельный ввод ВЗ-Н32#для не бронированного#кабеля, диаметром 18-25мм', 'Формат': 'А4', 'Кол.': 2, 'Примечание': None, 'Цена': 1188.38}, {'Обозначение': 'ВРПТ.685541.003', 'Наименование': 'Устройство заземления', 'Формат': None, 'Кол.': 1, 'Примечание': None, 'Цена': None}],
    'Стандартные изделия': [{'Обозначение': None, 'Наименование': 'Винт А2.М6-6gx10.019#ГОСТ 17473-80', 'Формат': 'А4', 'Кол.': 2, 'Примечание': None, 'Цена': None}, {'Обозначение': None, 'Наименование': 'Шайба 6 019 ГОСТ 6402-70', 'Формат': 'А4', 'Кол.': 2, 'Примечание': None, 'Цена': None}, {'Обозначение': None, 'Наименование': 'Шайба A.6.019 ГОСТ 11371-78', 'Формат': 'А4', 'Кол.': 2, 'Примечание': None, 'Цена': None}],
    'Детали': [{'Обозначение': 'ВРПТ.745551.005-240', 'Наименование': 'DIN-рейка NS35х7,5, L=240 мм', 'Формат': 'А4', 'Кол.': 1, 'Примечание': None, 'Цена': None}]}
    :return:
    '''

    return_dict_attribs = dict()

    tag_in_BOM_dxf = {'Формат': 'A', 'Зона': 'B', 'Поз.': 'C', 'Обозначение': 'D', 'Наименование': 'E', 'Кол.': 'F',
                      'Примечание': 'G', 'Цена': 'H'}

    start_row_int = 1
    startstart_row_int = 1

    for name_property in list_for_creating_BOM_with:
        start_row_int += 1
        startstart_row_int += 1
        return_dict_attribs[f'E{start_row_int}'] = name_property
        start_row_int += 2
        startstart_row_int += 2

        equipment_list = list_for_creating_BOM_with[name_property]
        for equip_dict in equipment_list:
            max_row = add_dict(dict_with_all_info_in_BOM_row=equip_dict,
                               count_row=startstart_row_int)
            for column_name in equip_dict:
                # if 'Цена' != column_name:
                for name in equip_dict[column_name]:
                    if column_name in tag_in_BOM_dxf:
                        tag_attrib = tag_in_BOM_dxf[column_name] + str(start_row_int)
                        return_dict_attribs[tag_attrib] = name
                        start_row_int += 1
                start_row_int = startstart_row_int
            start_row_int = max_row
            startstart_row_int = max_row

    return return_dict_attribs

def modify_dict_for_BOM(dict_all_attribs_for_bom):
    '''
    Нужно получить номер листа ключом, значение аттрибуты и их значения,
    :param dict_all_attribs_for_bom:{'E2': 'Сборочные единицы',
    'D4': 'ВРПТ.301172.024-021',
    'E4': 'Оболочка ВП.262512', 'A4': 'А4', 'F4': '1', 'G4': '', 'D5': 'ВРПТ.305311.001-025', 'D6': '', 'D7': '',
    'E5': 'Кабельный ввод ВЗ-Н25', 'E6': 'для не бронированного', 'E7': 'кабеля, диаметром 12-18мм', 'A5': 'А4',
    'A6': '', 'A7': '', 'F5': '6', 'F6': '', 'F7': '', 'G5': '', 'G6': '', 'G7': '', 'D8': 'ВРПТ.305311.001-016', 'D9': '', 'D10': '', 'E8': 'Кабельный ввод ВЗ-Н16', 'E9': 'для не бронированного', 'E10': 'кабеля, диаметром 3-8мм', 'A8': 'А4', 'A9': '', 'A10': '', 'F8': '3', 'F9': '', 'F10': '', 'G8': '', 'G9': '', 'G10': '', 'D11': 'ВРПТ.305311.001-012', 'D12': '', 'D13': '', 'E11': 'Кабельный ввод ВЗ-Н12', 'E12': 'для не бронированного', 'E13': 'кабеля, диаметром 2-6мм', 'A11': 'А4', 'A12': '', 'A13': '', 'F11': '2', 'F12': '', 'F13': '', 'G11': '', 'G12': '', 'G13': '', 'D14': 'ВРПТ.305311.001-032', 'D15': '', 'D16': '', 'E14': 'Кабельный ввод ВЗ-Н32', 'E15': 'для не бронированного', 'E16': 'кабеля, диаметром 18-25мм', 'A14': 'А4', 'A15': '', 'A16': '', 'F14': '2', 'F15': '', 'F16': '', 'G14': '', 'G15': '', 'G16': '', 'E18': 'Стандартные изделия', 'D20': '', 'D21': '', 'E20': 'Винт А2.М6-6gx10.019', 'E21': 'ГОСТ 17473-80', 'A20': 'А4', 'A21': '', 'F20': '2', 'F21': '', 'G20': '', 'G21': '', 'D22': '', 'E22': 'Шайба 6 019 ГОСТ 6402-70', 'A22': 'А4', 'F22': '2', 'G22': '', 'D23': '', 'E23': 'Шайба A.6.019 ГОСТ 11371-78', 'A23': 'А4', 'F23': '2', 'G23': '', 'E25': 'Детали', 'D27': 'ВРПТ.745551.005-240', 'E27': 'DIN-рейка NS35х7,5, L=240 мм', 'A27': 'А4', 'F27': '1', 'G27': '', 'E29': 'Прочие изделия', 'D31': 'TU16-2-GY', 'E31': 'Клемма проходная винтовая Iн=76А', 'A31': '', 'F31': '3', 'G31': ''}
    :return:
    '''

    return_dict = dict()

    for attrib_name in dict_all_attribs_for_bom:
        if (int(attrib_name[1:])-29) <= 0:
            if 1 not in return_dict:
                return_dict[1] = dict()
            return_dict[1][attrib_name] = dict_all_attribs_for_bom[attrib_name]
        else:
            page_number = ((int(attrib_name[1:]) - 29)//32) + 2
            if page_number not in return_dict:
                return_dict[page_number] = dict()

            return_dict[page_number][attrib_name[0] + str(int(attrib_name[1:]) - 29 - (32 * (page_number-2)))] = dict_all_attribs_for_bom[attrib_name]

    return return_dict





if __name__ == '__main__':

    a = {'E2': 'Сборочные единицы', 'D4': 'ВРПТ.301172.024-021', 'E4': 'Оболочка ВП.262512', 'A4': 'А4', 'F4': '1', 'G4': '', 'D5': 'ВРПТ.305311.001-025', 'D6': '', 'D7': '', 'E5': 'Кабельный ввод ВЗ-Н25', 'E6': 'для не бронированного', 'E7': 'кабеля, диаметром 12-18мм', 'A5': 'А4', 'A6': '', 'A7': '', 'F5': '6', 'F6': '', 'F7': '', 'G5': '', 'G6': '', 'G7': '', 'D8': 'ВРПТ.305311.001-016', 'D9': '', 'D10': '', 'E8': 'Кабельный ввод ВЗ-Н16', 'E9': 'для не бронированного', 'E10': 'кабеля, диаметром 3-8мм', 'A8': 'А4', 'A9': '', 'A10': '', 'F8': '3', 'F9': '', 'F10': '', 'G8': '', 'G9': '', 'G10': '', 'D11': 'ВРПТ.305311.001-012', 'D12': '', 'D13': '', 'E11': 'Кабельный ввод ВЗ-Н12', 'E12': 'для не бронированного', 'E13': 'кабеля, диаметром 2-6мм', 'A11': 'А4', 'A12': '', 'A13': '', 'F11': '2', 'F12': '', 'F13': '', 'G11': '', 'G12': '', 'G13': '', 'D14': 'ВРПТ.305311.001-032', 'D15': '', 'D16': '', 'E14': 'Кабельный ввод ВЗ-Н32', 'E15': 'для не бронированного', 'E16': 'кабеля, диаметром 18-25мм', 'A14': 'А4', 'A15': '', 'A16': '', 'F14': '2', 'F15': '', 'F16': '', 'G14': '', 'G15': '', 'G16': '', 'E18': 'Стандартные изделия', 'D20': '', 'D21': '', 'E20': 'Винт А2.М6-6gx10.019', 'E21': 'ГОСТ 17473-80', 'A20': 'А4', 'A21': '', 'F20': '2', 'F21': '', 'G20': '', 'G21': '', 'D22': '', 'E22': 'Шайба 6 019 ГОСТ 6402-70', 'A22': 'А4', 'F22': '2', 'G22': '', 'D23': '', 'E23': 'Шайба A.6.019 ГОСТ 11371-78', 'A23': 'А4', 'F23': '2', 'G23': '', 'E25': 'Детали', 'D27': 'ВРПТ.745551.005-240', 'E27': 'DIN-рейка NS35х7,5, L=240 мм', 'A27': 'А4', 'F27': '1', 'G27': '', 'E29': 'Прочие изделия', 'D31': 'TU16-2-GY', 'E31': 'Клемма проходная винтовая Iн=76А', 'A31': '', 'F31': '3', 'G31': '', 'D32': 'TU6-2-PE'}
    b = modify_dict_for_BOM(a)
    print(b)

