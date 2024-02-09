import datetime
import os

import ezdxf

import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import sys

from src._notuse_src.src.pdf_creator import pdf_main
from src._notuse_src.src.qt_creating.terminal_ui import TerminalPage

from src._notuse_src.src.dxf_creating import import_module, shell_create,terminal_create,inputs_create
from src._notuse_src.src.dxf_creating import move_inserts
from src._notuse_src.src.dxf_creating import dimension_create
from src._notuse_src.src.dxf_creating import border_create
from src._notuse_src.src.dxf_creating import BOM_create
from src._notuse_src.src.dxf_creating import cutside_works
from src._notuse_src.src.dxf_creating import nameplate
from src._notuse_src.src.price_directory import price_creator

class DxfCreator(TerminalPage):

    def __init__(self,
                 save_path = None,
                 path_to_csv = None,
                 path_to_dxf = None,
                 path_to_terminal_dxf = None,
                 path_to_verification_xlsx=None
                 ):

        '''БАЗА ПРИ ЗАПУСКЕ'''
        super().__init__(save_path=save_path,
                         path_to_csv=path_to_csv,
                         path_to_dxf=path_to_dxf,
                         path_to_terminal_dxf=path_to_terminal_dxf
                         )

        self.previewButton_leftMenu.clicked.connect(self.create_shell_dxf_after_selfkey)
        self.previewButton_leftMenu.clicked.connect(self.get_lwpolyline) # Переместить по хорошему в shellpage_ui
        self.previewButton_leftMenu.clicked.connect(self.get_coordinates_for_side)  # Переместить по хорошему в inputspage_ui
        self.previewButton_leftMenu.clicked.connect(self.create_inputs_dxf_after_shell)
        self.previewButton_leftMenu.clicked.connect(self.create_terminals_dxf_after_DIN_REYKA)
        self.previewButton_leftMenu.clicked.connect(self.create_border)
        self.previewButton_leftMenu.clicked.connect(self.create_dimension)
        self.previewButton_leftMenu.clicked.connect(self.write_table_border_a3_main)
        self.previewButton_leftMenu.clicked.connect(self.create_nameplate)
        self.previewButton_leftMenu.clicked.connect(self.save_doc_new)
        self.previewButton_leftMenu.clicked.connect(self.create_BOM)

        '''Для планшета ставлю значения спинбоксов заранее, т.к. не видно их и нет дизайнера тут'''
        # self.test_write()

        '''Заполнение options'''
        self.path_to_verification_xlsx = path_to_verification_xlsx
        self.create_dict_for_verification()
        self.write_rudes_name()
        self.write_rudes_data()

        '''Заполнение координат вставки блоков'''

    def set_logger(self,logger):
        self.logger = logger

    def create_shell_dxf_after_selfkey(self):
        '''Создание dxf оболочки'''
        if self.shell_key != None:
            if 'shell' in self.dict_for_save_blocks_before_draw.keys():
                self.doc_new = \
                    import_module.clear_base_doc_for_new(
                        dxfbase_path=self.path_to_dxf,
                        dict_for_save_blocks_before_draw=self.dict_for_save_blocks_before_draw)

                self.extreme_lines_in_all_blocks = shell_create.define_extreme_lines_in_all_blocks(doc=self.doc_new)

                self.topside_insert = \
                    shell_create.create_topside(doc=self.doc_new,
                                                shell_name=self.shell_name,
                                                list_name_added=self.list_added_blocks)

                self.topside_insert_extreme_lines = self.extreme_lines_in_all_blocks[self.shell_name+'_topside']
                self.topside_insert_extreme_lines['xy_0'] = (0.0,0.0)

                self.downside_insert = \
                    shell_create.create_downside(doc=self.doc_new,
                                                 shell_name=self.shell_name,
                                                 topside_extreme_lines=self.topside_insert_extreme_lines,
                                                 extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                                 list_name_added=self.list_added_blocks)

                self.downside_insert_extreme_lines = \
                    shell_create.define_downsinde_extreme_lines(
                                 extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                 downside_insert=self.downside_insert)

                self.upside_insert = \
                    shell_create.create_upside(doc=self.doc_new,
                                               shell_name=self.shell_name,
                                               topside_extreme_lines=self.topside_insert_extreme_lines,
                                               extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                               list_name_added=self.list_added_blocks)

                self.upside_insert_extreme_lines = \
                    shell_create.define_upside_extreme_lines(
                                 extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                 upside_insert=self.upside_insert)

                self.leftside_insert = \
                    shell_create.create_leftside(doc=self.doc_new,
                                                 shell_name=self.shell_name,
                                                 topside_extreme_lines=self.topside_insert_extreme_lines,
                                                 extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                                 list_name_added=self.list_added_blocks)

                self.leftside_insert_extreme_lines = \
                    shell_create.calculate_extreme_lines_in_leftside_insert(leftside_insert=self.leftside_insert)

                self.rightside_insert = \
                    shell_create.create_rightside(doc=self.doc_new,
                                                  shell_name=self.shell_name,
                                                  topside_extreme_lines=self.topside_insert_extreme_lines,
                                                  extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                                  list_name_added=self.list_added_blocks)
                self.right_insert_extreme_lines = \
                    shell_create.calculate_extreme_lines_in_rightside_insert(rightside_insert=self.rightside_insert)

                self.cutside_insert = \
                    shell_create.create_cutside_shell(doc=self.doc_new,
                                                      shell_name=self.shell_name,
                                                      leftside_extreme_lines=self.leftside_insert_extreme_lines,
                                                      extreme_line_all_blocks=self.extreme_lines_in_all_blocks,
                                                      list_name_added=self.list_added_blocks)

                self.cutside_insert_extreme_lines = \
                    shell_create.calculate_extreme_lines_in_cutside_insert(cutside_insert=self.cutside_insert)

                self.withoutcapside_insert = \
                    shell_create.create_withoutcapside_shell(doc=self.doc_new,
                                                             shell_name=self.shell_name,
                                                             cutside_extreme_lines=self.cutside_insert_extreme_lines,
                                                             extreme_line_in_all_blocks=self.extreme_lines_in_all_blocks,
                                                             list_name_added=self.list_added_blocks)
                self.withoutcapside_insert_extreme_lines = \
                    shell_create.calculate_extreme_lines_in_withoutcapside_insert(withoutcapside_insert=self.withoutcapside_insert)

                self.installation_dimension_insert = \
                    shell_create.create_installation_dimensions(doc=self.doc_new,
                                                                shell_name=self.shell_name,
                                                                extreme_lines_in_all_blocks=self.extreme_lines_in_all_blocks)

                self.din_insert = shell_create.create_din_reyka(doc=self.doc_new,
                                                                shell_name=self.shell_name,
                                                                list_name_added=self.list_added_blocks)


    def create_inputs_dxf_after_shell(self):
        '''Создание вводов на сторонах оболочки'''
        if hasattr(self,'doc_new'):
            '''Проверка на тип взрывозащиты'''
            ex_protection = None
            if 'exe' in self.safefactortypeCombobox_shellpage.currentText().lower():
                ex_protection = 'exe'
            elif 'exd' in self.safefactortypeCombobox_shellpage.currentText().lower():
                ex_protection = 'exd'
            else:
                ex_protection = None

            if ex_protection != None:
                inputs_create.create_inputs_in_block(doc=self.doc_new,
                                                     dict_before_match=self.dict_with_list_coordinates_on_side_for_dxf,
                                                     full_shale_name=self.shell_name,
                                                     type_of_explosion=ex_protection)

                self.input_max_len = move_inserts.move_shells_after_inputs(doc=self.doc_new,
                                                                           shell_name=self.shell_name)
                self.boundaries_drawing = move_inserts.get_boundaries_drawing(doc=self.doc_new,
                                                                              shell_name=self.shell_name,
                                                                              input_max_len=self.input_max_len)
                self.scale_drawing = move_inserts.define_scale(doc=self.doc_new,
                                                               shell_name=self.shell_name,
                                                               input_max_len=self.input_max_len)
                inputs_create.create_inputs_on_topside_withoutcapside(doc=self.doc_new,
                                                                      shell_name=self.shell_name)


    def create_terminals_dxf_after_DIN_REYKA(self):
        '''Добавление клемм на DIN рейку'''
        if hasattr(self,'doc_new'):
            din_reyka = terminal_create.check_din_reyka(self.doc_new, self.shell_name)
            if din_reyka:
                len_din_reyka = terminal_create.define_len_terminal(self.doc_new, din_reyka.dxf.name)
                if hasattr(self,'list_with_terminals'):
                    list_with_terminals = self.list_with_terminals

                    '''Поиск длины всех клемм и самой первой'''
                    summary_terminal_len = sum([terminal_create.define_len_terminal(self.doc_new,terminal)
                                                for terminal in list_with_terminals])

                    len_first_terminal = terminal_create.define_len_terminal(self.doc_new, list_with_terminals[0])

                    if len_din_reyka > summary_terminal_len:
                        terminal_create.create_terminal_on_din(doc_after_import=self.doc_new,
                                                               list_terminal_blocks=list_with_terminals,
                                                               din_reyka_insert=din_reyka)

                        terminal_create.create_terminal_on_cutside(doc_after_terminal=self.doc_new,
                                                                   list_terminal_blocks=list_with_terminals,
                                                                   shell_name=self.shell_name)

                    self.list_for_import_terminal = terminal_create.create_list_for_drawing_terminal(self.list_from_terminal_listwidget)

            move_inserts.scale_all_insert(doc=self.doc_new,
                                          scale=self.scale_drawing)

    def create_dimension(self):
        '''Создаем размер'''

        insert_topside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_topside"]')[0]

        extreme_lines_topside_after_scale = shell_create.define_extreme_lines_in_insert(insert=insert_topside)
        insert_upside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_upside"]')[0]
        extreme_lines_upside_after_scale = shell_create.define_extreme_lines_in_insert(insert=insert_upside)
        insert_downside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_downside"]')[0]
        extreme_lines_downside_after_scale = shell_create.define_extreme_lines_in_insert(insert=insert_downside)
        insert_rightside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_rightside"]')[0]
        extreme_lines_rightside_after_scale = shell_create.define_extreme_lines_in_insert(insert=insert_rightside)
        insert_on_topside = dimension_create.define_inputs_on_topside(doc=self.doc_new,
                                                                      shell_name=self.shell_name,
                                                                      extreme_lines_topside_after_scale=extreme_lines_topside_after_scale)

        point_for_horizontal_dimension = \
            {'max_up': dimension_create.calculate_max_up_coordinate(
                doc=self.doc_new,insert_on_side_dict=insert_on_topside, scale=self.scale_drawing,
                topside_extreme_lines=extreme_lines_topside_after_scale),
             'min_down': dimension_create.calculate_min_down_coordinate(
                 doc=self.doc_new, insert_on_side_dict=insert_on_topside, scale=self.scale_drawing,
                 topside_extreme_lines=extreme_lines_topside_after_scale),
             'min_left': dimension_create.calculate_min_left_coordinate(
                 doc=self.doc_new, insert_on_side_dict=insert_on_topside, scale=self.scale_drawing,
                 topside_extreme_lines=extreme_lines_topside_after_scale),
             'max_right':dimension_create.calculate_max_right_coordinate(
                 doc=self.doc_new, insert_on_side_dict=insert_on_topside, scale=self.scale_drawing,
                 topside_extreme_lines=extreme_lines_topside_after_scale)}

        # min_x_for_vertical_dim = min(point_for_horizontal_dimension['min_down'][0],
        #                              point_for_horizontal_dimension['max_up'][0])
        #
        # point_for_horizontal_dimension['min_down'][0] = min_x_for_vertical_dim
        # point_for_horizontal_dimension['max_up'][0] = min_x_for_vertical_dim

        dim = self.doc_new.modelspace().add_linear_dim(
            angle=90,
            p1=tuple(point_for_horizontal_dimension['min_down']),
            p2=tuple(point_for_horizontal_dimension['max_up']),
            dimstyle='EZDXF',
            base=(point_for_horizontal_dimension['min_left'][0] - self.input_max_len/self.scale_drawing,
                  point_for_horizontal_dimension['max_right'][0]- self.input_max_len/self.scale_drawing),
            text= f'{int((point_for_horizontal_dimension["max_up"][1] - point_for_horizontal_dimension["min_down"][1]) * self.scale_drawing)}'
        ).render()

        dim.dimension.dxf.text = f'{round(dim.dimension.get_measurement() * self.scale_drawing, 0)}'

        dim_horizontal = self.doc_new.modelspace().add_linear_dim(
            base=(point_for_horizontal_dimension['min_left'][0],
                  extreme_lines_topside_after_scale['y_min'] - 2.5 * self.input_max_len/self.scale_drawing),
                  #point_for_horizontal_dimension['min_down'][0] - 3 * self.input_max_len/self.scale_drawing),
            p1=point_for_horizontal_dimension['min_left'],
            p2=point_for_horizontal_dimension['max_right'],
            dimstyle='EZDXF')
        dim_horizontal.dimension.dxf.text = f'{int(dim_horizontal.dimension.get_measurement() * self.scale_drawing)}'
        dim_horizontal.render()

        dim_height_rightside = self.doc_new.modelspace().add_linear_dim(
            p1=(extreme_lines_rightside_after_scale['x_min'],extreme_lines_rightside_after_scale['y_max']),
            p2=(extreme_lines_rightside_after_scale['x_max'],extreme_lines_rightside_after_scale['y_max']),
            dimstyle='EZDXF',
            base=(extreme_lines_rightside_after_scale['x_min'] + (extreme_lines_rightside_after_scale['x_max'] - extreme_lines_rightside_after_scale['x_min'])/2,
                  extreme_lines_rightside_after_scale['y_max'] + 1.5 * self.input_max_len/self.scale_drawing)
        )

        dim_height_rightside.dimension.dxf.text = f'{int(dim_height_rightside.dimension.get_measurement() * self.scale_drawing)}'
        dim_height_rightside.set_arrows(blk=ezdxf.ARROWS.closed_filled)
        dim_height_rightside.render()

        cutside_works.add_label_cut(doc=self.doc_new,
                                    shell_name=self.shell_name,
                                    max_len_input=self.input_max_len / self.scale_drawing,
                                    scale=self.scale_drawing)

    def create_border(self):
        '''Создает рамку относительно '''
        insert_rightside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_rightside"]')[0]
        x_min = shell_create.define_extreme_lines_in_insert(insert=insert_rightside)['x_min']

        insert_upside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_upside"]')[0]
        y_min = shell_create.define_extreme_lines_in_insert(insert=insert_upside)['y_min']

        self.border_a3_insert = border_create.create_border_A3(doc=self.doc_new,
                                                          x_min_rightside=x_min,
                                                          y_min_upside=y_min)

        move_inserts.move_all_blocks_vertical_after_add_border(doc=self.doc_new,
                                                               shell_name=self.shell_name,
                                                               input_max_len=self.input_max_len/self.scale_drawing)
        self.create_nameplate_firstpage()
        self.create_nameplate_firstpage_withoutcapside()

    def test_write(self):
        self.manufactureComboboxWidget_shellpage.setCurrentText('ВЗОР')
        self.safefactortypeCombobox_shellpage.setCurrentText('Exe оболочки')
        self.serialCombobox_shellpage.setCurrentText('ВП')
        self.sizeCombobox_shellpage.setCurrentText('262512')
        self.gas_mark_RadioButton_shellpage.setChecked(True)
        self.gasdustoreComboBox_shellpage.setCurrentText('1Ex e IIC')
        self.temperature_class_comboBox_shellpage.setCurrentText('T4')
        self.manufacturerInputsComboBox.setCurrentText('ВЗОР')
        self.inputtypeComboBox.setCurrentText('ВЗ-Н')

        self.sideAListWidget.addItem('ВЗ-Р25')
        self.sideAListWidget.addItem('ВЗ-Р32')

        self.sideBListWidget.addItem('ВЗ-Н16')
        self.sideBListWidget.addItem('ВЗ-Н12')

        self.sideVListWidget.addItem('ВЗ-Н32')
        self.sideVListWidget.addItem('ВЗ-Н16')

        self.sideGListWidget.addItem('ВЗ-Н12')
        self.sideGListWidget.addItem('ВЗ-Н16')

        self.handwrite_inputspageComboBox.setCurrentText('ВЗ-Н25')
        self.siteASpinBox.setValue(2)
        self.siteBSpinBox.setValue(1)
        self.siteVSpinBox.setValue(2)
        self.manufacturer_terminal_combobox.setCurrentText('SUPU')
        self.mounttype_terminal_combobox.setCurrentText('Винтовые')
        self.appointment_terminal_combobox.setCurrentText('L')
        self.conductorsection_terminal_combobox.setCurrentText('16')
        self.count_terminal_spinbox.setValue(3)

    def create_dict_for_verification(self):
        '''
        Создает словарь для верификации, определяет имя и почту
        :return:
        '''
        self.dict_first_second_name = dict()
        if self.path_to_verification_xlsx !=None:
            wb = openpyxl.load_workbook(self.path_to_verification)
            ws = wb.active
            for cell in ws['F']:
                if cell.value:
                    if '@' in cell.value:
                        value = cell.value.split('@')[0]
                        full_name = ws[f'B{cell.row}'].value
                        first_name = full_name.split(' ')[0]
                        second_name = full_name.split(' ')[1]
                        third_name = None
                        if len(full_name.split(' ')) >= 3:
                            third_name = full_name.split(' ')[2]
                        if third_name:
                            self.dict_first_second_name[value] = first_name + f' {second_name[0]}.{third_name[0]}.'
                        else:
                            self.dict_first_second_name[value] = first_name + f' {second_name[0]}.'


    def write_rudes_name(self):
        '''
        Записываем кто разработал
        :return:
        '''

        computer_name_designer = os.getlogin()
        if computer_name_designer != '' and computer_name_designer != 'admin':
            if computer_name_designer in self.dict_first_second_name:
                self.rudesLineEdit.insert(self.dict_first_second_name[computer_name_designer])
            else:
                self.rudesLineEdit.insert('nooffice')
        else:
            self.rudesLineEdit.insert('admin')

    def write_rudes_data(self):
        '''
        Записываем дату, когда разработал
        :return:
        '''
        self.date_today = str(datetime.date.today())
        self.rudesdataLineEdit.insert(f'{self.date_today.split("-")[::-1][1]}.{self.date_today.split("-")[::-1][2]}')

    def write_table_border_a3_main(self):
        for attrib in self.border_a3_insert.attribs:
            if self.task_number != '':
                border_create.write_RUTITLE_attrib(
                    attrib_rutitle=attrib,
                    rutitle_text=f'К{self.serialCombobox_shellpage.currentText()}.'
                                 f'{self.sizeCombobox_shellpage.currentText()}.'
                                 f'{self.task_number}.{self.position_number} ВО')

            else:
                border_create.write_RUTITLE_attrib(
                    attrib_rutitle=attrib,
                    rutitle_text=f'К{self.serialCombobox_shellpage.currentText()}.'
                                 f'{self.sizeCombobox_shellpage.currentText()}.XXXX.XX ВО')
            border_create.write_rudes(attrib_rudes=attrib,rudes=self.designer_name)
            border_create.write_scale(attrib_SCALE=attrib,SCALE=self.scale_drawing)
            border_create.write_page_number(attrib_RUSHEET=attrib,sheet_number=1)
            border_create.write_page_numbers(attrib_RUSHTS=attrib,sheet_count=2)
            border_create.write_rudesdata(attrib_rudesdata=attrib,rudesdata=self.rudesdataLineEdit.text())

    def create_nameplate_firstpage(self):
        '''Создает nameplate посередине крышки
           Не проверяет, что на крышке что-то есть, в дальнейшем надо добавить проверку
        '''
        insert_topside = self.doc_new.modelspace().query(f'INSERT[name=="{self.shell_name}_topside"]')[0]

        block_nameplate = self.doc_new.blocks['nameplate']
        values = {attdef.dxf.tag: '' for attdef in block_nameplate.query('ATTDEF')}

        if self.doc_new.blocks.get('nameplate'):
            nameplate_docnew = self.doc_new.modelspace().add_blockref(name='nameplate',
                                                                  insert=insert_topside.dxf.insert,
                                                                  dxfattribs={
                                                                      'xscale': 1/(self.scale_drawing*2),
                                                                      'yscale': 1/(self.scale_drawing*2),
                                                                  })
            nameplate_docnew.add_auto_attribs(values)

        for attrib_in_nameplate in nameplate_docnew.attribs:
            if self.task_number != '':
                nameplate.write_attrib_box_full_name(attrib=attrib_in_nameplate,
                                                     full_name=f'К{self.serialCombobox_shellpage.currentText()}.{self.sizeCombobox_shellpage.currentText()}',
                                                     add_numbers=f'.{self.task_number}.{self.position_number}')
            else:
                nameplate.write_attrib_box_full_name(attrib=attrib_in_nameplate,
                                                     full_name=f'К{self.serialCombobox_shellpage.currentText()}.{self.sizeCombobox_shellpage.currentText()}')

            nameplate.write_explosion_tag(attrib=attrib_in_nameplate,
                                          gasdustore=self.gasdustoreComboBox_shellpage.currentText(),
                                          temperature_class=self.temperature_class_comboBox_shellpage.currentText())
            nameplate.write_minus_temperature(attrib=attrib_in_nameplate,
                                              minus_temperature=self.mintempLineEdit_shellpage.text())
            nameplate.write_plus_temperature(attrib=attrib_in_nameplate,
                                             plus_temperature=self.maxtempLineedit_shellpage.text())
            nameplate.write_voltage_current_frequency(attrib=attrib_in_nameplate)
            nameplate.write_batch_number(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_1(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_2(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_3(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_4(attrib=attrib_in_nameplate)

    def create_nameplate_firstpage_withoutcapside(self):
        '''Создает nameplate посередине крышки
           Не проверяет, что на крышке что-то есть, в дальнейшем надо добавить проверку
        '''
        block_withoutcapside = self.doc_new.blocks[f'{self.shell_name}_withoutcapside']

        block_nameplate = self.doc_new.blocks['nameplate']
        values = {attdef.dxf.tag: '' for attdef in block_nameplate.query('ATTDEF')}

        if self.doc_new.blocks.get('nameplate'):
            y_coordinate = self.extreme_lines_in_all_blocks[f'{self.shell_name}_withoutcapside']['y_max']-\
                           self.extreme_lines_in_all_blocks[f'nameplate']['y_max']/2 -\
                           15
            nameplate_docnew = block_withoutcapside.add_blockref(name='nameplate',
                                                                 insert=(0,y_coordinate),
                                                                 dxfattribs={
                                                                     'xscale': 1/2,
                                                                     'yscale': 1/2}
                                                                 )
            nameplate_docnew.add_auto_attribs(values)

        for attrib_in_nameplate in nameplate_docnew.attribs:
            if self.task_number != '':
                nameplate.write_attrib_box_full_name(attrib=attrib_in_nameplate,
                                                     full_name=f'К{self.serialCombobox_shellpage.currentText()}.{self.sizeCombobox_shellpage.currentText()}',
                                                     add_numbers=f'.{self.task_number}.{self.position_number}')
            else:
                nameplate.write_attrib_box_full_name(attrib=attrib_in_nameplate,
                                                     full_name=f'К{self.serialCombobox_shellpage.currentText()}.{self.sizeCombobox_shellpage.currentText()}')

            nameplate.write_explosion_tag(attrib=attrib_in_nameplate,
                                          gasdustore=self.gasdustoreComboBox_shellpage.currentText(),
                                          temperature_class=self.temperature_class_comboBox_shellpage.currentText())
            nameplate.write_minus_temperature(attrib=attrib_in_nameplate,
                                              minus_temperature=self.mintempLineEdit_shellpage.text())
            nameplate.write_plus_temperature(attrib=attrib_in_nameplate,
                                             plus_temperature=self.maxtempLineedit_shellpage.text())
            nameplate.write_voltage_current_frequency(attrib=attrib_in_nameplate)
            nameplate.write_batch_number(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_1(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_2(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_3(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_4(attrib=attrib_in_nameplate)

    def create_nameplate(self):
        '''Создает self.doc_nameplate'''
        self.doc_nameplate = nameplate.create_nameplate_doc(dxfbase_path=self.path_to_dxf)
        border_a3_insert = border_create.create_border_A3(doc=self.doc_nameplate,
                                                          x_min_rightside=0,
                                                          y_min_upside=0)

        extreme_lines_border = shell_create.define_extreme_lines_in_insert(insert=border_a3_insert)

        nameplate_insert = nameplate.create_nameplate_insert(doc_nameplate=self.doc_nameplate,
                                                             extreme_lines_border_insert=extreme_lines_border)


        insert_x = extreme_lines_border['x_min'] + \
                   (extreme_lines_border['x_max'] - extreme_lines_border['x_min']) / 2

        insert_y = extreme_lines_border['y_min'] + \
                   (extreme_lines_border['y_max'] - extreme_lines_border['y_min']) / 2

        nameplate_insert.dxf.insert = (insert_x,insert_y)

        extreme_lines_nameplate = shell_create.define_extreme_lines_in_insert(insert=nameplate_insert)

        for attrib_in_nameplate in nameplate_insert.attribs:
            if self.task_number != '':
                nameplate.write_attrib_box_full_name(attrib=attrib_in_nameplate,
                                                     full_name=f'К{self.serialCombobox_shellpage.currentText()}.{self.sizeCombobox_shellpage.currentText()}',
                                                     add_numbers=f'.{self.task_number}.{self.position_number}')
            else:
                nameplate.write_attrib_box_full_name(attrib=attrib_in_nameplate,
                                                     full_name=f'К{self.serialCombobox_shellpage.currentText()}.{self.sizeCombobox_shellpage.currentText()}')

            nameplate.write_explosion_tag(attrib=attrib_in_nameplate,
                                          gasdustore=self.gasdustoreComboBox_shellpage.currentText(),
                                          temperature_class=self.temperature_class_comboBox_shellpage.currentText())
            nameplate.write_minus_temperature(attrib=attrib_in_nameplate,
                                              minus_temperature=self.mintempLineEdit_shellpage.text())
            nameplate.write_plus_temperature(attrib=attrib_in_nameplate,
                                             plus_temperature=self.maxtempLineedit_shellpage.text())
            nameplate.write_voltage_current_frequency(attrib=attrib_in_nameplate)
            nameplate.write_batch_number(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_1(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_2(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_3(attrib=attrib_in_nameplate)
            nameplate.write_just_attrib_4(attrib=attrib_in_nameplate)


        for attrib in border_a3_insert.attribs:
            if attrib.dxf.tag == 'SCALE':
                attrib.dxf.text = f'2.5:1'
            if self.task_number != '':
                border_create.write_RUTITLE_attrib(
                    attrib_rutitle=attrib,
                    rutitle_text=f'К{self.serialCombobox_shellpage.currentText()}.'
                                 f'{self.sizeCombobox_shellpage.currentText()}.'
                                 f'{self.task_number}.{self.position_number} ВО')

            else:
                border_create.write_RUTITLE_attrib(
                    attrib_rutitle=attrib,
                    rutitle_text=f'К{self.serialCombobox_shellpage.currentText()}.'
                                 f'{self.sizeCombobox_shellpage.currentText()}.XXXX.XX ВО')
            border_create.write_rudes(attrib_rudes=attrib,rudes=self.designer_name)
            border_create.write_page_number(attrib_RUSHEET=attrib,sheet_number=2)
            border_create.write_page_numbers(attrib_RUSHTS=attrib,sheet_count=2)
            border_create.write_rudesdata(attrib_rudesdata=attrib,rudesdata=self.rudesdataLineEdit.text())

        dim_horizontal = self.doc_nameplate.modelspace().add_linear_dim(
            base=(extreme_lines_nameplate['x_min'],
                  extreme_lines_nameplate['y_max'] + 15),
            # point_for_horizontal_dimension['min_down'][0] - 3 * self.input_max_len/self.scale_drawing),
            p1=(extreme_lines_nameplate['x_min'],extreme_lines_nameplate['y_max']),
            p2=(extreme_lines_nameplate['x_max'],extreme_lines_nameplate['y_max']),
            dimstyle='EZDXF')

        dim_horizontal.dimension.dxf.text = f'58'
        dim_horizontal.dimension.dxf.color = 7
        dim_horizontal.dimstyle.dxf.dimtxt = 4
        dim_horizontal.dimstyle.dxf.dimasz = 4
        dim_horizontal.render()


    def create_BOM(self):
        '''Создаем dxf с BOM'''
        # path_to_xlsx = BOM_create.get_path_to_xlsx(main_window_class_instance=self)

        path_to_xlsx = '\\'.join(os.getcwd().split('\\')[0:]) + '\\src\\bd\\BOM\\bom.xlsx'

        # tag_in_BOM_dxf = {'Формат': 'A', 'Зона': 'B', 'Поз.': 'C', 'Обозначение': 'D', 'Наименование': 'E', 'Кол.': 'F',
        #                   'Примечание': 'G'}

        if path_to_xlsx:
            data_base_bom = BOM_create.read_BOM_base(xlsx_base_path=path_to_xlsx)

            dict_with_block_names_counts = BOM_create.create_dict_with_insert_names(doc=self.doc_new)

            list_for_creating_BOM = list()

            for block_name in dict_with_block_names_counts:
                for count_block, name_block_base in data_base_bom['Блок'].items():
                    if name_block_base == block_name:
                        _dict = {}
                        for _ in data_base_bom:
                            if _ != 'Блок':
                                _dict[_] = data_base_bom[_][count_block]
                        _dict['Кол.'] = dict_with_block_names_counts[block_name]
                        if 'Винт' in _dict['Наименование'] or 'Шайба' in _dict['Наименование']:
                            _dict['Кол.'] +=1
                        list_for_creating_BOM.append(_dict)

            list_for_creating_BOM_with = BOM_create.create_dict_main_properties(list_for_creating_BOM)

            dict_all_attribs_for_bom = BOM_create.dict_all_attrib_in_BOM(list_for_creating_BOM_with)

            last_dict_with_attribs = BOM_create.modify_dict_for_BOM(dict_all_attribs_for_bom)

            self.doc_bom = BOM_create.create_doc_BOM(dxfbase_path=self.path_to_dxf)


            if not hasattr(self, 'doc_filename_save_dxf'):
                self.doc_filename_save_dxf, _ = QtWidgets.QFileDialog.getSaveFileName(self,
                                                                                      directory='\\'.join(
                                                                                          os.getcwd().split('\\')[0:]),
                                                                                      caption="Сохранение BOM и xlsx файла",
                                                                                      filter='DXF Files(*.dxf)')

            for i in last_dict_with_attribs:
                if i == 1:
                    border_insert_first_page = BOM_create.create_BOM_FIRST(doc_bom=self.doc_bom)
                    dict_attribs = {attrib.dxf.tag: attrib for attrib in border_insert_first_page.attribs}
                    dict_attribs['SHEET_FIRST'].dxf.text = str(i)
                    dict_attribs['SHEET_COUNT'].dxf.text = str(max(list(last_dict_with_attribs.keys())))
                    dict_attribs['RUDES'].dxf.text = self.designer_name
                    dict_attribs['RUDESDATA'].dxf.text = f'{self.date_today.split("-")[::-1][1]}.{self.date_today.split("-")[::-1][2]}'
                    for attribs in last_dict_with_attribs[i]:
                        if attribs in list(dict_attribs.keys()):
                            dict_attribs[attribs].dxf.text = last_dict_with_attribs[i][attribs]
                    self.doc_bom.saveas(self.doc_filename_save_dxf[:-3] + f'_3.dxf')
                    pdf_main.save_pdf(self.doc_filename_save_dxf[:-3] + f'_3.dxf')
                else:
                    self.doc_bom.modelspace().delete_all_entities()
                    border_insert_second_page = BOM_create.create_BOM_SECOND(doc_bom=self.doc_bom)
                    dict_attribs = {attrib.dxf.tag: attrib for attrib in border_insert_second_page.attribs}
                    dict_attribs['SHEET_COUNT'].dxf.text = str(i)
                    for attribs in last_dict_with_attribs[i]:
                        if attribs in list(dict_attribs.keys()):
                            dict_attribs[attribs].dxf.text = last_dict_with_attribs[i][attribs]
                    self.doc_bom.saveas(self.doc_filename_save_dxf[:-3] + f'_{2+i}.dxf')
                    pdf_main.save_pdf(self.doc_filename_save_dxf[:-3] + f'_{2+i}.dxf')

            self.price_xlsx_wb = price_creator.result_xlsx(dict_columnrowname_value=dict_all_attribs_for_bom)
            self.price_xlsx_wb.save(self.doc_filename_save_dxf[:-3]+f'.xlsx')
        else:
            self.error_window.add_error('Не выбран путь для базы цен и оборудования xlsx')



if __name__ == '__main__':
    path_to_csv = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\bd'
    path_to_dxf_shell = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\dxf_base\\DXF_BASE.dxf'
    path_to_terminal_dxf = '\\'.join(os.getcwd().split('\\')[0:-1]) + '\\dxf_base\\DXF_BASE.dxf'

    app = QtWidgets.QApplication(sys.argv)
    welcome_window = DxfCreator(path_to_csv=path_to_csv,
                                    path_to_dxf = path_to_dxf_shell,
                                  path_to_terminal_dxf=path_to_terminal_dxf)
    welcome_window.show()
    sys.exit(app.exec_())