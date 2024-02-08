import os
import datetime
import ezdxf
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import sys

from src.qt_creating import terminal_ui

class OptionsPage(terminal_ui.TerminalPage):

    def __init__(self,
                 save_path = None,
                 path_to_csv = None,
                 path_to_dxf = None,
                 path_to_terminal_dxf = None,
                 path_to_verification_xlsx = None
                 ):

        '''БАЗА ПРИ ЗАПУСКЕ'''
        super().__init__(save_path=save_path,
                         path_to_csv=path_to_csv,
                         path_to_dxf=path_to_dxf,
                         path_to_terminal_dxf=path_to_terminal_dxf,
                         )

        self.path_to_verification_xlsx = path_to_verification_xlsx

        #Получение имени человека
        self.create_dict_for_verification()
        self.write_rudes_name()
        self.write_rudes_data()

    def create_dict_for_verification(self):
        '''
        Создает словарь для верификации, определяет имя и почту
        :return:
        '''
        wb = openpyxl.load_workbook(self.path_to_verification_xlsx)
        ws = wb.active
        self.dict_first_second_name = dict()
        for cell in ws['F']:
            if cell.value:
                if '@' in cell.value:
                    value = cell.value.split('@')[0]
                    full_name = ws[f'B{cell.row}'].value
                    first_name = full_name.split(' ')[0]
                    second_name = full_name.split(' ')[1]
                    third_name = None
                    if len(full_name.split(' ')) >= 3:
                        third_name = full_name.split(' ')[2]
                    if third_name:
                        self.dict_first_second_name[value] = first_name + f' {second_name[0]}.{third_name[0]}.'
                    else:
                        self.dict_first_second_name[value] = first_name + f' {second_name[0]}.'

    def write_rudes_name(self):
        '''
        Записываем кто разработал
        :return:
        '''

        if os.getlogin() != '' and os.getlogin() != 'admin':
            self.rudesLineEdit.insert(self.dict_first_second_name[os.getlogin()])

    def write_rudes_data(self):
        '''
        Записываем дату, когда разработал
        :return:
        '''
        self.date_today = str(datetime.date.today())
        self.rudesdataLineEdit.insert(f'{self.date_today.split("-")[::-1][1]}.{self.date_today.split("-")[::-1][2]}')





