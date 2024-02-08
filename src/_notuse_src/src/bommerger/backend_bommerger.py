#python -m PyQt5.uic.pyuic bommerger.ui -o frontend_bommerger.py

import sys

from PyQt5 import QtCore, QtGui, QtWidgets
import os
import frontend_bommerger
import openpyxl
import xlrd

class CustomFileDialogExcel(QtWidgets.QFileDialog):
    def __init__(self):
        super().__init__()

    def closeEvent(self,event):
        reply = QtWidgets.QMessageBox.question(self, 'Window Close', 'Are you sure you want to close the window?',
                                     QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)

        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()
            print('Window closed')
        else:
            event.ignore()

class ClssDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(ClssDialog, self).__init__(parent)

        self.verticalLayout = QtWidgets.QVBoxLayout(self)
        self.verticalLayout.setObjectName("verticalLayout")
        self.textLabel = QtWidgets.QLabel()
        self.textLabel.setObjectName("HelpLabel")
        self.verticalLayout.addWidget(self.textLabel)
        self.setWindowTitle("Help Window")
        self.textLabel.setText(" Самое важное правило, должны быть имена столбцов"
                               "\n Обозначение Наименование Количество"
                               "\n          Действия:"
                               "\n Нажать кнопку выбор папки, выбрать папку где все файлы для объединения"
                               "\n Слева список файлов, можно крестиком удалить ненужный"
                               "\n далее нажимаем объединить и файл падает в корневую папку с названием newnew")

    def btnClosed(self):
        self.close()


class Bommerger(QtWidgets.QMainWindow,frontend_bommerger.Ui_MainWindow):

    def __init__(self):
        '''БАЗА ПРИ ЗАПУСКЕ'''
        super().__init__()

        self.setupUi(self)
        self.excel_folder_dialog = CustomFileDialogExcel()
        self.dict_for_xlsx = {}


        '''Получение пути до папки c xlsx'''
        self.pushButton_2.clicked.connect(self.action_pushButton2)
        '''Удаление имени файла xlsx'''
        self.pushButton.clicked.connect(self.click_delete_xlsx_file)
        '''Объединение'''
        self.pushButton_3.clicked.connect(self.merge)
        '''Описание работы'''
        self.menubar.clear()
        self.action_help = QtWidgets.QAction('Help',self)
        self.menubar.addAction(self.action_help)
        self.action_help.triggered.connect(self.open_help_window)

        # self.menu.triggered.connect(self.open_help_window)
        # self.menubar.addAction(self.menu)


    def get_excel_folder(self):
        '''Выбор папки где нужно просуммировать excel'''
        self.excel_path = self.excel_folder_dialog.getExistingDirectory(self,
                                                             caption="Выбрать папку БД")
        if self.excel_path == '':
            delattr(self,'excel_path')


    def complete_listwidget(self):
        '''Заполнение названий файлов в listwidget'''
        if hasattr(self,'excel_path'):
            for file_name in os.listdir(self.excel_path):
                if file_name.endswith('xlsx') or file_name.endswith('xls') or file_name.endswith('xlsm'):
                    self.listWidget.addItem(file_name)


    def get_fullnames_excel(self):
        # pass
        self.list_fullnames_path = []
        self.list_fullnames_path_first = []
        for row_number in range(self.listWidget.count()):
            full_path_excel_file = self.excel_path + '//' + str(self.listWidget.item(row_number).text())
            self.list_fullnames_path_first.append(full_path_excel_file)
            self.list_fullnames_path.append(full_path_excel_file)
            # openpyxl.load_workbook(full_path_excel_file)

        print(self.list_fullnames_path)

    def create_xlsx(self,fullpathname: str = None):
        '''
        Создает файл xlsx в папке BOMs, если этот файл старого формата xls
        :param fullpathname: C:\\Users\\g.zubkov\\PycharmProjects\\BOMMerging\\КВН_18_Коробка_взрывозащищенная_из_нежавеющей_стали.xls
        :return:C:\\Users\\g.zubkov\\PycharmProjects\\BOMMerging\\КВН_18_Коробка_взрывозащищенная_из_нежавеющей_стали.xlsx
        '''
        if fullpathname:
            if fullpathname.endswith('xls'):
                # Открываем файл xls для чтения
                book = xlrd.open_workbook(fullpathname)
                sheet = book.sheet_by_index(0)

                # Создаем новый файл xlsx
                workbook = openpyxl.Workbook()
                sheet_new = workbook.active

                # Проходим по всем строкам и столбцам в файле xls и записываем данные в файл xlsx
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        sheet_new.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

                # Сохраняем новый файл xlsx
                workbook.save(fullpathname + 'x')
                return fullpathname + 'x'

            elif fullpathname.endswith('xlsx'):
                return fullpathname
            elif fullpathname.endswith('xlsm'):
                return fullpathname
            else:
                raise ValueError('Файл не xls и не xlsx')

    def create_correct_xlsx(self):
        for number_fullname, fullname_path in enumerate(self.list_fullnames_path.copy()):
            xlsx_name = self.create_xlsx(fullpathname=fullname_path)
            self.list_fullnames_path[number_fullname] = xlsx_name
        self.list_fullnames_path = list(set(self.list_fullnames_path))


    def return_correct_worksheet(self,workbook):
        '''Возвращает лист для двух типов икселей'''
        if len(list(workbook.sheetnames)) == 1:
            return workbook.active
        elif 'Спецификация' in list(workbook.sheetnames):
            return workbook['Спецификация']
        elif 'Для спецификации' in list(workbook.sheetnames):
            return workbook['Для спецификации']

    def search_name_str(self,worksheet):
        '''
        Получаем координаты нахождения ячейки Наименование и Количество
        :param worksheet:
        :return: {'Наименование': 'E1'...
        '''
        dict_with_cell_coordinates = dict()
        for row in worksheet.rows:
            for cell in row:
                if cell.value != None and cell.value != '':
                    if 'Наименование' in str(cell.value):
                        if 'Наименование' not in dict_with_cell_coordinates:
                            dict_with_cell_coordinates['Наименование'] = f'{cell.column_letter}{cell.row}'
                    if 'Количество' in str(cell.value) or 'Кол-во' in str(cell.value) or 'Колм' in str(cell.value):
                        if 'Количество' not in dict_with_cell_coordinates:
                            dict_with_cell_coordinates['Количество'] = f'{cell.column_letter}{cell.row}'
                    if 'Обозначение' in str(cell.value):
                        if 'Обозначение' not in dict_with_cell_coordinates:
                            dict_with_cell_coordinates['Обозначение'] = f'{cell.column_letter}{cell.row}'
        return dict_with_cell_coordinates

    def give_to_dict_name_and_count(self,dict_for_xlsx, dict_with_name_count_coordinates,worksheet):
        '''
        Суммирует по наименованию в словаре количество
        :param dict_for_xlsx:
        :param dict_with_name_count_coordinates: {'Наименование': 'E1'...
        :return:
        '''
        dict_for_xlsx_copy = dict_for_xlsx.copy()
        if dict_with_name_count_coordinates['Наименование'][1] == dict_with_name_count_coordinates['Количество'][1]:
            for cell in worksheet[dict_with_name_count_coordinates['Наименование'][0]]:
                if cell.value != None and \
                        worksheet[dict_with_name_count_coordinates['Количество'][0] + str(cell.row)].value != None:
                    if cell.value != '#N/A' and cell.value != '0' and cell.value != 0:
                        if cell.value not in dict_for_xlsx_copy:
                            try:
                                dict_for_xlsx_copy[cell.value] = {'Количество':
                                                                      float(worksheet[dict_with_name_count_coordinates[
                                                                                          'Количество'][0] + str(
                                                                          cell.row)].value),
                                                                  'Обозначение': worksheet[
                                                                      dict_with_name_count_coordinates['Обозначение'][
                                                                          0] + str(cell.row)].value}
                            except:
                                print(worksheet[dict_with_name_count_coordinates['Количество'][0] + str(cell.row)].value)
                        else:
                            dict_for_xlsx_copy[cell.value]['Количество'] += \
                                float(
                                    worksheet[dict_with_name_count_coordinates['Количество'][0] + str(cell.row)].value)
        return dict_for_xlsx_copy

    def delete_new_xlsx(self,list_first, list_second):
        '''После создания xlsx удаляем их'''
        for file_name in list_second:
            if file_name not in list_first:
                os.remove(f'{file_name}')

    def create_new_xlsx(self,dict_for_xlsx_after_all):
        '''
        Создаем xlsx по полученным данным
        :param dict_for_xlsx_after_all:
        :param file_name: 'check.xlsx'
        :return:
        '''
        wb = openpyxl.Workbook()
        ws = wb.active
        current_row = 2
        ws['A1'] = 'Обозначение'
        ws['B1'] = 'Наименование'
        ws['C1'] = 'Количество'

        for name, count in dict_for_xlsx_after_all.items():
            ws.cell(row=current_row, column=2).value = name
            ws.cell(row=current_row, column=3).value = count['Количество']
            ws.cell(row=current_row, column=1).value = count['Обозначение']
            current_row += 1
        return wb


    def merge(self):
        if self.listWidget.count() != 0:
            self.get_fullnames_excel()
        if hasattr(self,'list_fullnames_path'):
            self.create_correct_xlsx()
        for xlsx_fullnamepath in sorted(self.list_fullnames_path):
            workbook = openpyxl.load_workbook(xlsx_fullnamepath, data_only=True)
            worksheet = self.return_correct_worksheet(workbook=workbook)
            dict_with_cell_coordinates = self.search_name_str(worksheet=worksheet)
            self.dict_for_xlsx = self.give_to_dict_name_and_count\
                (dict_for_xlsx=self.dict_for_xlsx,
                 dict_with_name_count_coordinates=dict_with_cell_coordinates,
                 worksheet=worksheet)
        self.dict_for_xlsx = dict(sorted(self.dict_for_xlsx.items()))
        self.delete_new_xlsx(list_first=self.list_fullnames_path_first,
                            list_second=self.list_fullnames_path)
        workbook_new = self.create_new_xlsx(dict_for_xlsx_after_all=self.dict_for_xlsx)
        workbook_new.save('newnew.xlsx')





    def action_pushButton2(self):
        '''Все действия при нажатии кнопки выбора папки'''
        self.listWidget.clear()
        self.get_excel_folder()
        self.complete_listwidget()
        self.get_fullnames_excel()


    def click_delete_xlsx_file(self):
        '''Удаление файлов xlsx которые не нужны из списка'''
        if hasattr(self,'excel_path'):
            rowIndex = self.listWidget.currentRow()
            currentItem = self.listWidget.takeItem(rowIndex)

    def open_help_window(self):
        '''Открытие окна help'''
        help_window = ClssDialog(self)
        help_window.exec_()


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    bommerger_window = Bommerger()
    bommerger_window.show()
    sys.exit(app.exec_())