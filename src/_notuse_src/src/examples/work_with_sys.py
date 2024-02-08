import sys
import os
import openpyxl
import ezdxf
import ezdxf
# from ezdxf.tools import fonts
from src.dxf_creating import BOM_create
from ezdxf.fonts import fonts

# # xdg_path() returns "$XDG_CACHE_HOME/ezdxf" or "~/.cache/ezdxf" if
# # $XDG_CACHE_HOME is not set
# font_cache_dir = ezdxf.options.xdg_path("XDG_CACHE_HOME", ".cache")
# fonts.build_system_font_cache(path=font_cache_dir)
# ezdxf.options.font_cache_directory = font_cache_dir
# # Save changes to the default config file "~/.config/ezdxf/ezdxf.ini"
# # to load the font cache always from the new location.
# ezdxf.options.write_home_config()
ezdxf.fonts.fonts.load()
# fonts.build_system_font_cache()
def result_xlsx():
    data= {'E2': 'Сборочные единицы', 'D4': 'ВРПТ.301172.024-015', 'E4': 'Оболочка ВП.261610', 'A4': 'А4', 'F4': '1', 'G4': '', 'D5': 'ВРПТ.305311.002-020', 'D6': '', 'D7': '', 'D8': '', 'E5': 'Кабельный ввод ВЗ-Б20', 'E6': 'для бронированного кабеля,', 'E7': 'наружным диаметром 9-17мм,', 'E8': 'внутренним диаметром 6-14мм', 'A5': 'А4', 'A6': '', 'A7': '', 'A8': '', 'F5': '5', 'F6': '', 'F7': '', 'F8': '', 'G5': '', 'G6': '', 'G7': '', 'G8': '', 'D9': 'ВРПТ.685541.003', 'E9': 'Устройство заземления', 'A9': '', 'F9': '1', 'G9': '', 'E11': 'Стандартные изделия', 'D13': '', 'D14': '', 'E13': 'Винт А2.М6-6gx10.019', 'E14': 'ГОСТ 17473-80', 'A13': 'А4', 'A14': '', 'F13': '2', 'F14': '', 'G13': '', 'G14': '', 'D15': '', 'E15': 'Шайба 6 019 ГОСТ 6402-70', 'A15': 'А4', 'F15': '2', 'G15': '', 'D16': '', 'E16': 'Шайба A.6.019 ГОСТ 11371-78', 'A16': 'А4', 'F16': '2', 'G16': ''}

    from openpyxl import Workbook

    # Создание нового XLSX файла
    wb = Workbook()

    # Получение активного листа
    sheet1 = wb.active
    sheet1.title = "Прайс по позициям"

    # Заполнение первого листа данными из словаря

    for cell, value in data.items():
        sheet1[cell] = value

    # Создание второго листа
    sheet2 = wb.create_sheet(title="Лист 2")

    # Заполнение второго листа данными из словаря
    row = 1
    for cell, value in data.items():
        if cell.startswith('D'):
            sheet2.cell(row=row, column=1, value=value)
            row += 1

    # Сохранение файла
    wb.save("C:\\Users\\g.zubkov\\PycharmProjects\\Project_not_for_exe\\example.xlsx")

something = {'Сборочные единицы': [{'Обозначение': ['ВРПТ.301172.024-10 '], 'Наименование': ['Оболочка ВП.122209'], 'Формат': ['А4'], 'Кол.': ['1'], 'Примечание': [''], 'Цена': ['1954.54']}, {'Обозначение': ['ВРПТ.305311.002-016', '', '', ''], 'Наименование': ['Кабельный ввод ВЗ-Б16', 'для бронированного кабеля,', 'наружным диаметром 8-12мм,', 'внутренним диаметром 3-8мм'], 'Формат': ['А4', '', '', ''], 'Кол.': ['3', '', '', ''], 'Примечание': ['', '', '', ''], 'Цена': ['800', '', '', '']}], 'Стандартные изделия': [{'Обозначение': ['', ''], 'Наименование': ['Винт А2.М6-6gx10.019', 'ГОСТ 17473-80'], 'Формат': ['А4', ''], 'Кол.': ['2', ''], 'Примечание': ['', ''], 'Цена': ['', '']}, {'Обозначение': [''], 'Наименование': ['Шайба 6 019 ГОСТ 6402-70'], 'Формат': ['А4'], 'Кол.': ['2'], 'Примечание': [''], 'Цена': ['']}, {'Обозначение': [''], 'Наименование': ['Шайба A.6.019 ГОСТ 11371-78'], 'Формат': ['А4'], 'Кол.': ['2'], 'Примечание': [''], 'Цена': ['']}], 'Детали': [{'Обозначение': ['ВРПТ.745551.005-180'], 'Наименование': ['DIN-рейка NS35х7,5, L=180 мм'], 'Формат': ['А4'], 'Кол.': ['1'], 'Примечание': [''], 'Цена': ['']}]}

print(BOM_create.dict_all_attrib_in_BOM(something))

