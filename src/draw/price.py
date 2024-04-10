import openpyxl
from openpyxl import Workbook
import math
def read_price(path_price):
    price_wb = openpyxl.load_workbook(path_price, data_only=True)
    price_terminal_ws = price_wb['4. РС клеммы']
    price_terminal_dict = dict()
    for cell in price_terminal_ws['C']:
        if isinstance(price_terminal_ws[f'I{cell.row}'].value,float) or isinstance(price_terminal_ws[f'I{cell.row}'].value,int):
            price_terminal_dict[cell.value] = math.ceil(price_terminal_ws[f'I{cell.row}'].value)
            # print(type(price_terminal_ws[f'I{cell.row}'].value))

    return price_terminal_dict

def result_xlsx(dict_columnrowname_value:dict,price_terminal_dict:dict):
    '''
    Получение экселя с ценами
    :param dict_columnrowname_value:     data= {'E2': 'Сборочные единицы', 'D4': 'ВРПТ.301172.024-015', 'E4': 'Оболочка ВП.261610', 'A4': 'А4', 'F4': '1', 'G4': '', 'D5': 'ВРПТ.305311.002-020', 'D6': '', 'D7': '', 'D8': '', 'E5': 'Кабельный ввод ВЗ-Б20', 'E6': 'для бронированного кабеля,', 'E7': 'наружным диаметром 9-17мм,', 'E8': 'внутренним диаметром 6-14мм', 'A5': 'А4', 'A6': '', 'A7': '', 'A8': '', 'F5': '5', 'F6': '', 'F7': '', 'F8': '', 'G5': '', 'G6': '', 'G7': '', 'G8': '', 'D9': 'ВРПТ.685541.003', 'E9': 'Устройство заземления', 'A9': '', 'F9': '1', 'G9': '', 'E11': 'Стандартные изделия', 'D13': '', 'D14': '', 'E13': 'Винт А2.М6-6gx10.019', 'E14': 'ГОСТ 17473-80', 'A13': 'А4', 'A14': '', 'F13': '2', 'F14': '', 'G13': '', 'G14': '', 'D15': '', 'E15': 'Шайба 6 019 ГОСТ 6402-70', 'A15': 'А4', 'F15': '2', 'G15': '', 'D16': '', 'E16': 'Шайба A.6.019 ГОСТ 11371-78', 'A16': 'А4', 'F16': '2', 'G16': ''}
    :return:
    '''
    # Создание нового XLSX файла
    wb = Workbook()

    # Получение активного листа
    sheet1 = wb.active
    sheet1.title = "Прайс по позициям"

    # Заполнение первого листа данными из словаря

    for cell, value in dict_columnrowname_value.copy().items():
        if value in price_terminal_dict:
            dict_columnrowname_value[f'H{cell[1:]}'] = str(price_terminal_dict[value])


    for cell, value in dict_columnrowname_value.items():
        sheet1[cell] = value

    # Создание второго листа
    sheet2 = wb.create_sheet(title="Текущие цены на позиции")

    # Заполнение второго листа данными из словаря

    row = 1
    for cell, value in dict_columnrowname_value.items():
        if cell.startswith('E'):
            sheet2.cell(row=int(cell[1:]), column=1, value=value)
        elif cell.startswith('H'):
            sheet2.cell(row=int(cell[1:]), column=2, value=value)

        # row +=1

    # Сохранение файла
    return wb

if __name__ == '__main__':
    print(read_price(path_price='C:\\Users\\g.zubkov\\PycharmProjects\\SpecmashSPR\\price.xlsx'))